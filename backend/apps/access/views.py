"""
apps/access/views.py

ViewSets / APIViews:
  - AulaViewSet            : CRUD for doors
  - ScheduleViewSet        : CRUD for time windows
  - AccessPermissionViewSet: CRUD for user-door-schedule grants
  - AccessEventViewSet     : Read-only audit log
  - AccessValidateView     : POST /api/access/validate/ (Raspberry Pi endpoint)
  - KPIView                : GET  /api/access/kpi/
  - ReporteView            : GET  /api/access/reports/summary/
"""

from rest_framework import viewsets, status, mixins, filters
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework.views import APIView
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.utils import extend_schema, extend_schema_view, OpenApiParameter, OpenApiResponse
from drf_spectacular.types import OpenApiTypes

from django.utils import timezone
from django.db.models import Count, Q, Avg
from django.db.models.functions import TruncHour, TruncDate
from django.http import HttpResponse
from datetime import datetime, timedelta
import csv
import random
import string

from .models import Aula, Schedule, AccessPermission, AccessEvent, TeacherOTP
from .serializers import (
    AulaSerializer,
    ScheduleSerializer,
    AccessPermissionSerializer,
    AccessEventSerializer,
    AccessValidateSerializer,
    TeacherOTPSerializer,
)
from .services import AccessService, AccessValidationInput, AccessMethod
from .filters import AccessEventFilter
from apps.users.models import User as LocalUser


# ─────────────────────────────────────────
# Aulas
# ─────────────────────────────────────────

@extend_schema_view(
    list=extend_schema(
        tags=["Access"],
        summary="Listar aulas",
        description=(
            "Retorna todas las aulas/puertas del sistema.\n\n"
            "**Filtros:** `is_active`, `desired_state`, `actual_state`\n"
            "**Búsqueda:** `search` busca por código o descripción"
        ),
        parameters=[
            OpenApiParameter("search", OpenApiTypes.STR, description="Busca por código o descripción del aula"),
            OpenApiParameter("is_active", OpenApiTypes.BOOL, description="Filtra aulas activas/inactivas"),
            OpenApiParameter("desired_state", OpenApiTypes.STR, description="Estado deseado de la puerta"),
            OpenApiParameter("actual_state", OpenApiTypes.STR, description="Estado actual de la puerta"),
        ],
    ),
    create=extend_schema(
        tags=["Access"],
        summary="Crear aula",
        description="Registra una nueva aula/puerta en el sistema.\n\n**Campos requeridos:** `code`, `description`",
    ),
    retrieve=extend_schema(tags=["Access"], summary="Obtener aula por ID"),
    update=extend_schema(tags=["Access"], summary="Actualizar aula (completo)"),
    partial_update=extend_schema(tags=["Access"], summary="Actualizar aula (parcial)"),
    destroy=extend_schema(tags=["Access"], summary="Eliminar aula"),
)
class AulaViewSet(viewsets.ModelViewSet):
    """CRUD for physical doors/classrooms."""

    queryset = Aula.objects.all().order_by("code")
    serializer_class = AulaSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ["is_active", "desired_state", "actual_state"]
    search_fields = ["code", "description"]


# ─────────────────────────────────────────
# Horarios (Schedule)
# ─────────────────────────────────────────

@extend_schema_view(
    list=extend_schema(
        tags=["Access"],
        summary="Listar horarios",
        description=(
            "Retorna todos los bloques horarios del sistema.\n\n"
            "**Filtro:** `day_of_week` (0=Lunes … 6=Domingo)"
        ),
        parameters=[
            OpenApiParameter(
                "day_of_week", OpenApiTypes.INT,
                description="Día de la semana: 0=Lunes, 1=Martes, …, 6=Domingo"
            ),
        ],
    ),
    create=extend_schema(
        tags=["Access"],
        summary="Crear horario",
        description=(
            "Registra un bloque horario.\n\n"
            "**Campos requeridos:** `day_of_week` (0–6), `start_time` (HH:MM), `end_time` (HH:MM)"
        ),
    ),
    retrieve=extend_schema(tags=["Access"], summary="Obtener horario por ID"),
    update=extend_schema(tags=["Access"], summary="Actualizar horario (completo)"),
    partial_update=extend_schema(tags=["Access"], summary="Actualizar horario (parcial)"),
    destroy=extend_schema(tags=["Access"], summary="Eliminar horario"),
)
class ScheduleViewSet(viewsets.ModelViewSet):
    """CRUD for weekly schedule time windows."""

    queryset = Schedule.objects.all().order_by("day_of_week", "start_time")
    serializer_class = ScheduleSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["day_of_week"]


# ─────────────────────────────────────────
# Permisos de acceso
# ─────────────────────────────────────────

@extend_schema_view(
    list=extend_schema(
        tags=["Access"],
        summary="Listar permisos de acceso",
        description=(
            "Retorna todas las asignaciones usuario→aula→horario.\n\n"
            "Incluye campos enriquecidos: `user_nombre`, `user_email`, `aula_code`, `schedule_display`.\n\n"
            "**Filtros:** `user` (UUID), `aula` (UUID), `is_active`"
        ),
        parameters=[
            OpenApiParameter("user", OpenApiTypes.UUID, description="UUID del usuario"),
            OpenApiParameter("aula", OpenApiTypes.UUID, description="UUID del aula"),
            OpenApiParameter("is_active", OpenApiTypes.BOOL, description="Solo permisos activos/inactivos"),
        ],
    ),
    create=extend_schema(
        tags=["Access"],
        summary="Crear permiso de acceso",
        description=(
            "Otorga acceso a un usuario para un aula en un horario específico.\n\n"
            "**Body:** `{ \"user\": \"<uuid>\", \"aula\": \"<uuid>\", \"schedule\": \"<uuid>\", \"is_active\": true }`"
        ),
    ),
    retrieve=extend_schema(tags=["Access"], summary="Obtener permiso por ID"),
    update=extend_schema(tags=["Access"], summary="Actualizar permiso (completo)"),
    partial_update=extend_schema(tags=["Access"], summary="Actualizar permiso (parcial)"),
    destroy=extend_schema(tags=["Access"], summary="Revocar permiso de acceso"),
)
class AccessPermissionViewSet(viewsets.ModelViewSet):
    """CRUD for user access grants (user → aula → schedule)."""

    queryset = AccessPermission.objects.select_related("user", "aula", "schedule").all()
    serializer_class = AccessPermissionSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["user", "aula", "is_active"]

    @extend_schema(
        summary="Upsert de permiso desde el calendario",
        description="Busca/crea un horario (Schedule) y vincula un permiso de acceso en una sola operación.",
        request=OpenApiTypes.OBJECT, # Simplificado para el ejemplo
        responses={200: AccessPermissionSerializer, 201: AccessPermissionSerializer}
    )
    @action(detail=False, methods=["post"])
    def upsert_calendar_event(self, request):
        user_id = request.data.get("user")
        aula_id = request.data.get("aula")
        day_of_week = request.data.get("day_of_week")
        start_time = request.data.get("start_time")
        end_time = request.data.get("end_time")
        is_anytime = request.data.get("is_anytime", False)
        is_recurring = request.data.get("is_recurring", True)
        date = request.data.get("date")
        permission_id = request.data.get("id")

        if not user_id or not aula_id:
            return Response({"error": "user y aula son requeridos"}, status=400)

        # 1. Buscar o Crear el Schedule
        schedule, _ = Schedule.objects.get_or_create(
            day_of_week=day_of_week if is_recurring else None,
            date=date if not is_recurring else None,
            start_time=start_time,
            end_time=end_time,
            is_recurring=is_recurring,
            is_anytime=is_anytime
        )

        # 2. Upsert Permission
        if permission_id:
            try:
                permission = AccessPermission.objects.get(id=permission_id)
                permission.user_id = user_id
                permission.aula_id = aula_id
                permission.schedule = schedule
                permission.is_active = True
                permission.save()
            except AccessPermission.DoesNotExist:
                return Response({"error": "Permiso no encontrado"}, status=404)
        else:
            permission = AccessPermission.objects.create(
                user_id=user_id,
                aula_id=aula_id,
                schedule=schedule,
                is_active=True
            )

        serializer = self.get_serializer(permission)
        return Response(serializer.data, status=status.HTTP_201_CREATED if not permission_id else status.HTTP_200_OK)


# ─────────────────────────────────────────
# Eventos de acceso (read-only audit log)
# ─────────────────────────────────────────

@extend_schema_view(
    list=extend_schema(
        tags=["Access"],
        summary="Listar eventos de acceso",
        description=(
            "Retorna el registro de auditoría de todos los accesos al sistema. **Solo lectura.**\n\n"
            "Ordenado por timestamp descendente (más reciente primero).\n\n"
            "**Filtros:** `method` (FACE/PIN/MANUAL), `result` (SUCCESS/DENIED), `alert_flag`, `user`, `aula`, `device`\n"
            "**Ordering:** `timestamp`"
        ),
        parameters=[
            OpenApiParameter("method", OpenApiTypes.STR, description="Método: FACE, PIN o MANUAL"),
            OpenApiParameter("result", OpenApiTypes.STR, description="Resultado: SUCCESS o DENIED"),
            OpenApiParameter("alert_flag", OpenApiTypes.BOOL, description="Solo eventos marcados como alerta"),
            OpenApiParameter("user", OpenApiTypes.UUID, description="UUID del usuario"),
            OpenApiParameter("aula", OpenApiTypes.UUID, description="UUID del aula"),
            OpenApiParameter("ordering", OpenApiTypes.STR, description="Ordena por: timestamp, -timestamp"),
        ],
    ),
    retrieve=extend_schema(
        tags=["Access"],
        summary="Obtener evento por ID",
        description="Retorna el detalle de un evento de acceso específico.",
    ),
)
class AccessEventViewSet(
    mixins.ListModelMixin,
    mixins.RetrieveModelMixin,
    viewsets.GenericViewSet,
):
    """Read-only audit log. Creation is done exclusively by AccessService."""

    queryset = AccessEvent.objects.select_related("user", "aula", "device").all()
    serializer_class = AccessEventSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = AccessEventFilter
    ordering_fields = ["timestamp"]
    ordering = ["-timestamp"]

    @action(detail=False, methods=["get"])
    def export_csv(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="reporte_accesos.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Fecha', 'Usuario', 'Email', 'Aula', 'Método', 'Resultado', 'Alerta', 'Razón'])
        
        for event in queryset:
            writer.writerow([
                event.id,
                event.timestamp,
                event.user.full_name if event.user else 'Anónimo',
                event.user.email if event.user else 'N/A',
                event.aula.code if event.aula else 'N/A',
                event.method,
                event.result,
                'SÍ' if event.alert_flag else 'NO',
                event.reason or ''
            ])
        
        return response

    @action(detail=False, methods=["get"])
    def export_excel(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        queryset = self.filter_queryset(self.get_queryset())
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Accesos"
        
        headers = ['ID', 'Fecha', 'Usuario', 'Email', 'Aula', 'Método', 'Resultado', 'Alerta', 'Razón']
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            
        for event in queryset:
            ws.append([
                str(event.id),
                event.timestamp.replace(tzinfo=None) if event.timestamp else '',
                event.user.full_name if event.user else 'Anónimo',
                event.user.email if event.user else 'N/A',
                event.aula.code if event.aula else 'N/A',
                event.method,
                event.result,
                'SÍ' if event.alert_flag else 'NO',
                event.reason or ''
            ])
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="reporte_accesos.xlsx"'
        wb.save(response)
        return response


# ─────────────────────────────────────────
# Validación de acceso (Raspberry Pi)
# ─────────────────────────────────────────

@extend_schema(
    tags=["Access"],
    summary="Validar intento de acceso",
    description=(
        "**Endpoint principal del dispositivo Raspberry Pi.**\n\n"
        "Recibe un intento de acceso biométrico o PIN y retorna si el acceso es permitido o denegado.\n\n"
        "**Métodos:** `FACE` (facial), `PIN` (contingencia), `MANUAL` (apertura manual)\n\n"
    ),
    request=AccessValidateSerializer,
    responses={
        200: OpenApiResponse(description="Acceso procesado — incluye `result`, `correlation_id`, `event_id`"),
        400: OpenApiResponse(description="Datos de entrada inválidos"),
    },
)
class AccessValidateView(APIView):
    """POST /api/access/validate/ — Used by Raspberry Pi devices."""

    permission_classes = [IsAuthenticated]

    def post(self, request: Request) -> Response:
        serializer = AccessValidateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        validated = serializer.validated_data

        try:
            # Source device_id from the lock associated with the aula
            from apps.devices.models import Lock
            lock = Lock.objects.filter(aula_id=validated["aula_id"]).first()
            device_id = lock.device_id if lock else None

            payload = AccessValidationInput(
                method=AccessMethod(validated["method"]),
                data=validated["data"],
                aula_id=validated["aula_id"],
                device_id=device_id,
            )
            result = AccessService.validate(payload)
            return Response(
                {
                    "result": result.result.value,
                    "correlation_id": str(result.correlation_id),
                    "event_id": str(result.event_id),
                    "reason": result.reason,
                    "user_full_name": result.user_full_name,
                },
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            import traceback
            traceback.print_exc()
            logger.error(f"Access validation failed: {e}", exc_info=True)
            return Response({"error": str(e)}, status=500)


# ─────────────────────────────────────────
# KPI Dashboard
# ─────────────────────────────────────────

@extend_schema(
    tags=["Access"],
    summary="KPIs del dashboard",
    description=(
        "Retorna métricas en tiempo real calculadas desde los eventos de acceso del día actual."
    ),
    responses={200: OpenApiResponse(description="KPIs calculados exitosamente")},
)
class KPIView(APIView):
    """GET /api/access/kpi/ — Real-time dashboard metrics."""

    permission_classes = [IsAuthenticated]

    def get(self, request: Request) -> Response:
        start_str = request.query_params.get('start_date')
        end_str = request.query_params.get('end_date')

        # 1. Period Calculation
        try:
            if start_str and end_str:
                start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
            else:
                start_date = timezone.localdate()
                end_date = start_date
        except ValueError:
            return Response({"error": "Formato de fecha inválido. Use YYYY-MM-DD."}, status=400)

        delta = (end_date - start_date).days + 1
        prev_start = start_date - timedelta(days=delta)
        prev_end = end_date - timedelta(days=delta)

        # 2. Query Helper
        def get_stats(s_date, e_date):
            qs = AccessEvent.objects.filter(timestamp__date__range=(s_date, e_date))
            return qs.aggregate(
                total=Count("id"),
                success=Count("id", filter=Q(result="SUCCESS")),
                denied=Count("id", filter=Q(result="DENIED")),
                falsos_negativos=Count("id", filter=Q(method="FACE", result="DENIED")),
                uso_otp=Count("id", filter=Q(method="OTP")),
                score_promedio=Avg("score", filter=Q(method="FACE", score__isnull=False)),
                tiempo_respuesta_promedio=Avg("response_time"),
                alertas_activas=Count("id", filter=Q(alert_flag=True))
            )

        curr_stats = get_stats(start_date, end_date)
        prev_stats = get_stats(prev_start, prev_end)

        # 3. Trend Calculation Helper
        def calc_trend(curr, prev):
            if not prev: return {"value": 0, "isPositive": True} if not curr else {"value": 100, "isPositive": True}
            diff = ((curr - prev) / prev) * 100
            return {"value": round(abs(diff), 1), "isPositive": diff >= 0}

        def calc_inverse_trend(curr, prev): # For "bad" metrics like rejection or response time
            trend = calc_trend(curr, prev)
            # Inverting isPositive doesn't make sense for the value, but for UI color. 
            # We'll just return the raw trend and let the UI decide if "Positive" is Green or Red.
            return trend

        # 4. Chart Data (Dynamic Granularity)
        events_curr = AccessEvent.objects.filter(timestamp__date__range=(start_date, end_date))
        
        if delta <= 1:
            # Group by Hour if single day
            by_time = (
                events_curr
                .annotate(period=TruncHour("timestamp"))
                .values("period")
                .annotate(cantidad=Count("id"))
                .order_by("period")
            )
            chart_data = [
                {"hora": f"{item['period'].hour:02d}:00", "cantidad": item["cantidad"]}
                for item in by_time if item["period"]
            ]
            chart_key = "accesos_por_hora"
        else:
            # Group by Date if multiple days
            by_time = (
                events_curr
                .annotate(period=TruncDate("timestamp"))
                .values("period")
                .annotate(cantidad=Count("id"))
                .order_by("period")
            )
            chart_data = [
                {"hora": item['period'].strftime('%d %b'), "cantidad": item["cantidad"]}
                for item in by_time if item["period"]
            ]
            chart_key = "accesos_por_dia"

        top_aulas = (
            events_curr
            .values("aula__code")
            .annotate(cantidad=Count("id"))
            .order_by("-cantidad")[:5]
        )

        por_metodo = (
            events_curr
            .values("method")
            .annotate(cantidad=Count("id"))
            .order_by("-cantidad")
        )

        total = curr_stats["total"] or 0
        prev_total = prev_stats["total"] or 0
        success = curr_stats["success"] or 0
        prev_success = prev_stats["success"] or 0
        
        tasa_exito = (success / total * 100) if total else 0
        prev_tasa_exito = (prev_success / prev_total * 100) if prev_total else 0

        return Response({
            "total_accesos": total,
            "total_accesos_trend": calc_trend(total, prev_total),
            "tasa_exito": round(tasa_exito, 1),
            "tasa_exito_trend": calc_trend(tasa_exito, prev_tasa_exito),
            "tasa_rechazo": round(100 - tasa_exito, 1) if total else 0,
            "tasa_rechazo_trend": calc_trend(100 - tasa_exito, 100 - prev_tasa_exito) if total and prev_total else {"value":0, "isPositive":True},
            "falsos_negativos": curr_stats["falsos_negativos"] or 0,
            "falsos_negativos_trend": calc_trend(curr_stats["falsos_negativos"] or 0, prev_stats["falsos_negativos"] or 0),
            "uso_otp": curr_stats["uso_otp"] or 0,
            "uso_otp_trend": calc_trend(curr_stats["uso_otp"] or 0, prev_stats["uso_otp"] or 0),
            "score_promedio": round(curr_stats["score_promedio"] or 0, 1),
            "score_promedio_trend": calc_trend(curr_stats["score_promedio"] or 0, prev_stats["score_promedio"] or 0),
            "tiempo_respuesta_promedio": round(curr_stats["tiempo_respuesta_promedio"] or 0, 3),
            "tiempo_respuesta_trend": calc_trend(curr_stats["tiempo_respuesta_promedio"] or 0, prev_stats["tiempo_respuesta_promedio"] or 0),
            "alertas_activas": curr_stats["alertas_activas"] or 0,
            "alertas_activas_trend": calc_trend(curr_stats["alertas_activas"] or 0, prev_stats["alertas_activas"] or 0),
            "usuarios_activos": LocalUser.objects.filter(is_active=True).count(),
            chart_key: chart_data,
            "top_aulas": [
                {"aula": item["aula__code"] or "N/A", "cantidad": item["cantidad"]}
                for item in top_aulas
            ],
            "accesos_por_metodo": [
                {"metodo": item["method"], "cantidad": item["cantidad"]}
                for item in por_metodo
            ],
            "start_date": str(start_date),
            "end_date": str(end_date),
            "is_today": start_date == timezone.localdate() and end_date == timezone.localdate()
        })


# ─────────────────────────────────────────
# Reporte mensual / Personalizado
# ─────────────────────────────────────────

@extend_schema(
    tags=["Access"],
    summary="Resumen de reportes (rango de fechas)",
    description=(
        "Genera un resumen estadístico de los accesos en un rango de fechas:\n\n"
        "- `start_date`: Fecha inicio (YYYY-MM-DD)\n"
        "- `end_date`: Fecha fin (YYYY-MM-DD)\n"
    ),
    parameters=[
        OpenApiParameter("start_date", OpenApiTypes.DATE, description="Fecha de inicio"),
        OpenApiParameter("end_date", OpenApiTypes.DATE, description="Fecha de fin"),
    ],
    responses={200: OpenApiResponse(description="Reporte generado exitosamente")},
)
class ReporteView(APIView):
    """GET /api/access/reports/summary/ — Date-range based report."""

    permission_classes = [IsAuthenticated]

    def get(self, request: Request) -> Response:
        start_str = request.query_params.get('start_date')
        end_str = request.query_params.get('end_date')

        if start_str and end_str:
            try:
                start = timezone.make_aware(datetime.strptime(start_str, '%Y-%m-%d'))
                end = timezone.make_aware(datetime.strptime(end_str, '%Y-%m-%d')) + timedelta(days=1)
            except ValueError:
                return Response({"error": "Formato de fecha inválido. Use YYYY-MM-DD."}, status=400)
        else:
            end = timezone.now()
            start = end - timedelta(days=30)

        events = AccessEvent.objects.filter(timestamp__range=(start, end))
        total = events.count()
        permitidos = events.filter(result="SUCCESS").count()
        denegados = events.filter(result="DENIED").count()

        por_dia = (
            events.annotate(fecha=TruncDate("timestamp"))
            .values("fecha")
            .annotate(
                permitidos=Count("id", filter=Q(result="SUCCESS")),
                denegados=Count("id", filter=Q(result="DENIED")),
            )
            .order_by("fecha")
        )

        por_metodo = events.values("method").annotate(cantidad=Count("id"))

        # Most active users in the period
        usuarios_activos = (
            events.filter(user__isnull=False)
            .values("user__nombre", "user__apellido", "user__email")
            .annotate(cantidad=Count("id"))
            .order_by("-cantidad")[:5]
        )

        tasa_puntualidad = round((permitidos / total * 100), 1) if total else 0

        return Response({
            "periodo": f"{start.date()} - {end.date() - timedelta(days=1)}",
            "total_accesos": total,
            "accesos_permitidos": permitidos,
            "accesos_denegados": denegados,
            "tasa_puntualidad": tasa_puntualidad,
            "accesos_por_dia": [
                {"fecha": str(item["fecha"]), "permitidos": item["permitidos"], "denegados": item["denegados"]}
                for item in por_dia
            ],
            "accesos_por_metodo": [
                {"metodo": item["method"], "cantidad": item["cantidad"]}
                for item in por_metodo
            ],
            "heatmap": [],
            "usuarios_mas_activos": [
                {
                    "nombre": f"{u['user__nombre']} {u['user__apellido']}".strip() or u["user__email"],
                    "cantidad": u["cantidad"]
                }
                for u in usuarios_activos
            ],
        })



class TeacherOTPViewSet(
    mixins.CreateModelMixin,
    mixins.ListModelMixin,
    viewsets.GenericViewSet
):
    """
    ViewSet for managing Teacher OTPs.
    Teachers can generate their own codes here.
    """
    queryset = TeacherOTP.objects.all()
    serializer_class = TeacherOTPSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # Users (Teachers) can only see their own OTPs
        # Map auth.User to LocalUser via email
        auth_user = self.request.user
        local_user = LocalUser.objects.filter(email=auth_user.email).first()
        if not local_user:
            return self.queryset.none()
        return self.queryset.filter(user=local_user)

    @action(detail=False, methods=["post"])
    def generate(self, request):
        auth_user = request.user
        # Find the local user profile linked by email
        local_user = LocalUser.objects.filter(email=auth_user.email).first()
        
        if not local_user:
            return Response(
                {"error": "Su cuenta no tiene un perfil de docente vinculado (email no encontrado)."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Invalidate existing unused OTPs for this user to keep it clean
        TeacherOTP.objects.filter(user=local_user, is_used=False).update(is_used=True)
        
        # Generate random 6-digit code
        code = "".join(random.choices(string.digits, k=6))
        expires_at = timezone.now() + timedelta(minutes=10)
        
        otp = TeacherOTP.objects.create(
            user=local_user,
            code=code,
            expires_at=expires_at
        )
        
        data = TeacherOTPSerializer(otp).data
        data["message"] = "Código OTP generado exitosamente. Válido por 10 minutos."
        return Response(data, status=status.HTTP_201_CREATED)
